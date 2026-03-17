"""
NSE1 Consolidator
=================
Merges all country Excel files into one master file.

Usage:
    python consolidate.py

Setup:
    1. Put all country Excel files in the 'country_files' folder
    2. Run this script — it creates/updates 'NSE1 Important Dates.xlsx'

Each country file must have columns: Country, Date, Event, Status, Description
"""

COUNTRY_FOLDER = "country_files"
MASTER_FILE    = "NSE1 Important Dates.xlsx"

import sys, os, glob
from datetime import datetime

def check_dependencies():
    missing = []
    try: import pandas
    except ImportError: missing.append("pandas")
    try: import openpyxl
    except ImportError: missing.append("openpyxl")
    if missing:
        print(f"\n❌  Missing: pip install {' '.join(missing)}\n")
        sys.exit(1)

check_dependencies()
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

REQUIRED_COLS = {'Country', 'Date', 'Event', 'Status', 'Description'}

def load_country_files():
    if not os.path.exists(COUNTRY_FOLDER):
        print(f"\n❌  Folder '{COUNTRY_FOLDER}' not found.")
        print(f"    Create it and put country Excel files inside.\n")
        sys.exit(1)

    files = glob.glob(os.path.join(COUNTRY_FOLDER, "*.xlsx"))
    if not files:
        print(f"\n❌  No .xlsx files found in '{COUNTRY_FOLDER}'\n")
        sys.exit(1)

    frames, errors = [], []
    for f in files:
        name = os.path.basename(f)
        try:
            df = pd.read_excel(f)
            missing = REQUIRED_COLS - set(df.columns)
            if missing:
                errors.append(f"  ⚠️  {name} — missing columns: {missing}")
                continue
            df['_source'] = name
            frames.append(df)
            print(f"  ✅  {name} — {len(df)} rows")
        except Exception as e:
            errors.append(f"  ❌  {name} — error: {e}")

    if errors:
        print("\nWarnings:")
        for e in errors: print(e)

    if not frames:
        print("\n❌  No valid files loaded.\n")
        sys.exit(1)

    return pd.concat(frames, ignore_index=True)

def validate_and_clean(df):
    today = pd.Timestamp.now().normalize()

    df['Date']        = pd.to_datetime(df['Date'], errors='coerce')
    df['Description'] = df['Description'].fillna('').astype(str).str.strip()
    df['Status']      = df['Status'].fillna('').astype(str).str.strip()
    df['Event']       = df['Event'].fillna('').astype(str).str.strip()
    df['Country']     = df['Country'].fillna('').astype(str).str.strip()

    df = df.dropna(subset=['Date'])
    df = df[df['Event'] != '']

    issues = []

    # Rule 1: Executed but date is in future → revert
    future_executed = (df['Status'].str.lower() == 'executed') & (df['Date'] > today)
    if future_executed.any():
        issues.append(f"  ⚠️  {future_executed.sum()} event(s) marked Executed but date hasn't passed → reverted to 'Yet to Happen'")
        df.loc[future_executed, 'Status'] = 'Yet to Happen'

    # Rule 2: Hide descriptions for future events
    future_with_desc = (df['Date'] > today) & (df['Description'] != '')
    if future_with_desc.any():
        issues.append(f"  ℹ️   {future_with_desc.sum()} future event(s) had descriptions → hidden until date passes")
        df.loc[future_with_desc, 'Description'] = ''

    if issues:
        print("\nValidation:")
        for i in issues: print(i)

    return df

def save_master(df):
    out = df[['Country', 'Date', 'Event', 'Status', 'Description']].copy()
    out = out.sort_values(['Date', 'Country']).reset_index(drop=True)
    out['Date'] = out['Date'].dt.strftime('%d-%b-%Y')
    out.to_excel(MASTER_FILE, index=False)

    wb = load_workbook(MASTER_FILE)
    ws = wb.active

    header_fill = PatternFill("solid", start_color="1B3F8B", end_color="1B3F8B")
    header_font = Font(bold=True, color="FFFFFF", name="Arial", size=10)
    thin   = Side(style="thin", color="DDDDDD")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal="center", vertical="center")
    left   = Alignment(horizontal="left",   vertical="center", wrap_text=True)

    for cell in ws[1]:
        cell.fill = header_fill; cell.font = header_font
        cell.alignment = center; cell.border = border

    exec_fill = PatternFill("solid", start_color="E8F5EE", end_color="E8F5EE")
    post_fill = PatternFill("solid", start_color="FFF3E8", end_color="FFF3E8")
    yet_fill  = PatternFill("solid", start_color="E6ECF8", end_color="E6ECF8")

    for row in ws.iter_rows(min_row=2):
        status = str(row[3].value or '').lower()
        fill = exec_fill if 'executed' in status else post_fill if 'postponed' in status else yet_fill
        for cell in row:
            cell.fill = fill; cell.border = border
            cell.font = Font(name="Arial", size=9)
            cell.alignment = left if cell.column in [3, 5] else center

    ws.column_dimensions['A'].width = 16
    ws.column_dimensions['B'].width = 14
    ws.column_dimensions['C'].width = 38
    ws.column_dimensions['D'].width = 16
    ws.column_dimensions['E'].width = 50
    ws.row_dimensions[1].height = 20
    ws.freeze_panes = "A2"
    wb.save(MASTER_FILE)

def main():
    print("\n── NSE1 Consolidator ───────────────────────────────")
    print(f"  Scanning: {COUNTRY_FOLDER}/\n")

    df = load_country_files()
    df = validate_and_clean(df)
    save_master(df)

    total     = len(df)
    countries = df['Country'].nunique()
    executed  = df['Status'].str.lower().str.contains('executed').sum()
    yet       = df['Status'].str.lower().str.contains('yet').sum()
    with_desc = (df['Description'] != '').sum()

    print(f"\n  Countries: {countries}  |  Events: {total}")
    print(f"  Executed: {executed}  |  Upcoming: {yet}  |  With reports: {with_desc}")
    print(f"  Saved:    {MASTER_FILE}")
    print("────────────────────────────────────────────────────")
    print("  ✅  Master file ready! Now run update_dashboard.py")
    print("────────────────────────────────────────────────────\n")

if __name__ == "__main__":
    main()
